# Copyright (c) 2026 Anders Ødenes. All rights reserved.
"""Import lat/lon coordinates from FieldFlow v1 into v2 locations.

Sources:
  1. Årskontroller Excel files (sheet "Alle_jobber") — finds "ticket number", "lat", "lon"
     columns by name. Matches jobs.external_id → location_id → updates locations.
  2. geocode_results.json (address|postal_code → [lat, lon])
     Matches on locations.address + locations.postal_code
  3. Google Maps Geocoding API — for remaining locations without coordinates.
     Requires GOOGLE_MAPS_API_KEY env var. Use --geocode flag to enable.
"""
import asyncio
import json
import os
import sys
import time
from pathlib import Path

import requests

# Add project root to path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from dotenv import load_dotenv

load_dotenv(PROJECT_ROOT / ".env", override=True)
DATABASE_URL = os.environ["DATABASE_URL"].replace("postgresql+asyncpg://", "postgresql://")

try:
    import openpyxl
except ImportError:
    print("openpyxl er ikke installert. Kjør: pip install openpyxl")
    sys.exit(1)

import asyncpg

# ── FieldFlow v1 data directory ───────────────────────────────────────────────
V1_DATA_DIR = Path(
    r"C:\Users\anders.odenes\OneDrive - Oy Hedengren Ab"
    r"\Työpöytä\FieldFlow\data"
)

# Region → main Excel file
REGION_FILES = {
    "Oslo": V1_DATA_DIR / "oslo" / "Årskontroller Oslo.xlsx",
    "Bergen": V1_DATA_DIR / "bergen" / "Årskontroller Bergen.xlsx",
    "Stavanger": V1_DATA_DIR / "stavanger" / "Årskontroller Stavanger.xlsx",
}

SHEET_NAME = "Alle_jobber"

# Column names to find (case-insensitive)
TICKET_COL_NAMES = ["ticket number", "ticket_number"]
LAT_COL_NAMES = ["lat", "latitude"]
LON_COL_NAMES = ["lon", "longitude"]

# Region → geocode_results.json (only Bergen has one)
GEOCODE_FILES = {
    "Bergen": V1_DATA_DIR / "bergen" / "geocode_results.json",
}


def find_col_by_name(headers: list[str], candidates: list[str]) -> int | None:
    """Find column index by matching header name (case-insensitive)."""
    for i, h in enumerate(headers):
        if h and h.strip().lower() in candidates:
            return i
    return None


def read_excel_coordinates(filepath: Path) -> dict[str, tuple[float, float]]:
    """Read ticket_number → (lat, lon) from the 'Alle_jobber' sheet.

    Finds columns by name, not by index, since column layout varies per region.
    Returns dict of {ticket_number: (lat, lon)} for rows with valid coordinates.
    """
    coords: dict[str, tuple[float, float]] = {}

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"  ADVARSEL: Ark '{SHEET_NAME}' ikke funnet. Tilgjengelige: {wb.sheetnames}")
        wb.close()
        return coords

    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(values_only=True)

    # Read header row and find columns by name
    header = next(rows_iter, None)
    if not header:
        wb.close()
        return coords

    headers = [str(h).strip().lower() if h else "" for h in header]

    ticket_idx = find_col_by_name(headers, TICKET_COL_NAMES)
    lat_idx = find_col_by_name(headers, LAT_COL_NAMES)
    lon_idx = find_col_by_name(headers, LON_COL_NAMES)

    if ticket_idx is None or lat_idx is None or lon_idx is None:
        missing = []
        if ticket_idx is None:
            missing.append("ticket number")
        if lat_idx is None:
            missing.append("lat")
        if lon_idx is None:
            missing.append("lon")
        print(f"  ADVARSEL: Fant ikke kolonner: {missing}")
        print(f"  Kolonner i arket: {headers}")
        wb.close()
        return coords

    print(f"  Kolonner: ticket={headers[ticket_idx]} (idx {ticket_idx}), "
          f"lat={headers[lat_idx]} (idx {lat_idx}), lon={headers[lon_idx]} (idx {lon_idx})")

    max_idx = max(ticket_idx, lat_idx, lon_idx)
    no_coords = 0
    for row in rows_iter:
        if len(row) <= max_idx:
            continue

        ticket_val = row[ticket_idx]
        lat_val = row[lat_idx]
        lon_val = row[lon_idx]

        if not ticket_val:
            continue

        ticket = str(ticket_val).strip()
        if not ticket:
            continue

        try:
            lat = float(lat_val)
            lon = float(lon_val)
        except (TypeError, ValueError):
            no_coords += 1
            continue

        # Sanity check: Norwegian coordinates (lat ~57-72, lon ~4-32)
        if not (55 <= lat <= 73 and 3 <= lon <= 33):
            no_coords += 1
            continue

        coords[ticket] = (lat, lon)

    if no_coords:
        print(f"  {no_coords} rader uten gyldige koordinater")

    wb.close()
    return coords


def read_geocode_results(filepath: Path) -> dict[str, tuple[float, float]]:
    """Read geocode_results.json → dict of {"address|postal_code": (lat, lon)}."""
    coords: dict[str, tuple[float, float]] = {}

    with open(filepath, encoding="utf-8") as f:
        data = json.load(f)

    results = data.get("results", {})
    skipped = 0

    for key, latlon in results.items():
        if not isinstance(latlon, list) or len(latlon) != 2:
            skipped += 1
            continue

        try:
            lat = float(latlon[0])
            lon = float(latlon[1])
        except (TypeError, ValueError):
            skipped += 1
            continue

        if not (55 <= lat <= 73 and 3 <= lon <= 33):
            skipped += 1
            continue

        coords[key] = (lat, lon)

    failed_count = len(data.get("failed", []))
    if skipped:
        print(f"  Hoppet over {skipped} ugyldige oppslag")
    if failed_count:
        print(f"  {failed_count} mislykkede geokodinger i filen")

    return coords


def normalize_address(address: str) -> str:
    """Normalize address for matching: lowercase, strip, collapse whitespace."""
    return " ".join(address.lower().strip().split())


# ── Google Maps Geocoding ─────────────────────────────────────────────────────

def geocode_address(address: str, postal_code: str, api_key: str) -> tuple[float, float] | None:
    """Geocode a single address via Google Maps Geocoding API.

    Returns (lat, lon) or None if geocoding fails.
    """
    full_address = f"{address}, {postal_code}, Norge"
    url = "https://maps.googleapis.com/maps/api/geocode/json"
    params = {"address": full_address, "key": api_key}

    try:
        resp = requests.get(url, params=params, timeout=10)
        data = resp.json()

        if data["status"] == "OK":
            loc = data["results"][0]["geometry"]["location"]
            lat, lon = loc["lat"], loc["lng"]
            # Sanity check
            if 55 <= lat <= 73 and 3 <= lon <= 33:
                return (lat, lon)
            return None
        elif data["status"] == "OVER_QUERY_LIMIT":
            print("    API-grense nadd, venter 60s...")
            time.sleep(60)
            return geocode_address(address, postal_code, api_key)
        else:
            return None
    except Exception:
        return None


async def import_coordinates():
    do_geocode = "--geocode" in sys.argv
    dry_run = "--dry-run" in sys.argv

    print(f"Database: {DATABASE_URL[:40]}...")
    if do_geocode:
        api_key = os.environ.get("GOOGLE_MAPS_API_KEY", "")
        if not api_key:
            print("FEIL: GOOGLE_MAPS_API_KEY ikke satt i .env")
            print("Legg til: GOOGLE_MAPS_API_KEY=AIza...")
            return
        print(f"Google Maps API: aktivert (key: {api_key[:10]}...)")
    if dry_run:
        print("DRY RUN: ingen endringer i databasen")
    print()

    # ── Source 1: Excel files (ticket number → lat/lon) ───────────────────────
    print("=== Kilde 1: Excel-filer (ticket number → lat/lon) ===")
    excel_coords: dict[str, tuple[float, float]] = {}
    excel_counts: dict[str, int] = {}

    for region_name, filepath in REGION_FILES.items():
        print(f"\nLeser {region_name}: {filepath.name}")
        if not filepath.exists():
            print(f"  FIL IKKE FUNNET: {filepath}")
            continue

        coords = read_excel_coordinates(filepath)
        excel_counts[region_name] = len(coords)
        print(f"  {len(coords)} rader med koordinater")
        excel_coords.update(coords)

    print(f"\nTotalt {len(excel_coords)} unike ticket numbers med koordinater fra Excel")
    print()

    # ── Source 2: geocode_results.json (address|postal_code → lat/lon) ────────
    print("=== Kilde 2: geocode_results.json (adresse matching) ===")
    geocode_coords: dict[str, tuple[float, float]] = {}
    geocode_counts: dict[str, int] = {}

    for region_name, filepath in GEOCODE_FILES.items():
        print(f"\nLeser {region_name}: {filepath.name}")
        if not filepath.exists():
            print(f"  FIL IKKE FUNNET: {filepath}")
            continue

        coords = read_geocode_results(filepath)
        geocode_counts[region_name] = len(coords)
        print(f"  {len(coords)} adresser med koordinater")
        geocode_coords.update(coords)

    print(f"\nTotalt {len(geocode_coords)} adresseoppslag fra geocode_results")
    print()

    # ── Connect to database ──────────────────────────────────────────────────
    conn = await asyncpg.connect(DATABASE_URL, statement_cache_size=0)
    try:
        # Find tenant
        row = await conn.fetchrow("SELECT id FROM tenants WHERE slug = $1", "hedengren")
        if not row:
            print("Tenant 'hedengren' ikke funnet. Kjor seed.py forst.")
            return
        tenant_id = row["id"]
        print(f"Tenant: hedengren (id: {tenant_id})")

        # Set RLS context
        await conn.execute(f"SET app.current_tenant = '{tenant_id}'")

        # Load jobs with external_id → service_contract_id
        db_jobs = await conn.fetch(
            "SELECT j.external_id, sc.location_id "
            "FROM jobs j "
            "JOIN service_contracts sc ON j.service_contract_id = sc.id "
            "WHERE j.tenant_id = $1 AND j.external_id IS NOT NULL",
            tenant_id,
        )
        ticket_to_location: dict[str, object] = {}
        for r in db_jobs:
            if r["external_id"] not in ticket_to_location:
                ticket_to_location[r["external_id"]] = r["location_id"]

        print(f"Jobber i v2-database med external_id: {len(db_jobs)}")
        print(f"Unike ticket -> location mappings: {len(ticket_to_location)}")

        # Load all locations
        db_locations = await conn.fetch(
            "SELECT id, address, postal_code, latitude, longitude "
            "FROM locations WHERE tenant_id = $1",
            tenant_id,
        )
        loc_by_id = {r["id"]: r for r in db_locations}
        print(f"Lokasjoner i v2-database: {len(db_locations)}")
        print()

        # ── Pass 1: Match via ticket number → location_id ────────────────────
        updated_by_ticket = 0
        already_set_ticket = 0
        ticket_no_match = 0
        updated_location_ids: set = set()

        for ticket, (lat, lon) in excel_coords.items():
            loc_id = ticket_to_location.get(ticket)
            if loc_id is None:
                ticket_no_match += 1
                continue

            loc_row = loc_by_id.get(loc_id)
            if loc_row is None:
                ticket_no_match += 1
                continue

            if loc_row["latitude"] is not None and loc_row["longitude"] is not None:
                if (abs(loc_row["latitude"] - lat) < 0.0001
                        and abs(loc_row["longitude"] - lon) < 0.0001):
                    already_set_ticket += 1
                    updated_location_ids.add(loc_id)
                    continue

            if not dry_run:
                await conn.execute(
                    "UPDATE locations SET latitude = $1, longitude = $2 WHERE id = $3",
                    lat, lon, loc_id,
                )
            updated_by_ticket += 1
            updated_location_ids.add(loc_id)

        print(f"Pass 1 (ticket number): {updated_by_ticket} oppdatert, "
              f"{already_set_ticket} allerede satt, {ticket_no_match} uten match")

        # ── Pass 2: Match via address|postal_code (geocode_results) ──────────
        updated_by_addr = 0
        already_set_addr = 0
        geocode_no_match = 0

        db_by_addr: dict[str, list] = {}
        for loc_row in db_locations:
            if loc_row["id"] in updated_location_ids:
                continue
            addr = loc_row["address"] or ""
            postal = loc_row["postal_code"] or ""
            key = f"{normalize_address(addr)}|{postal.strip()}"
            db_by_addr.setdefault(key, []).append(loc_row)

        for geocode_key, (lat, lon) in geocode_coords.items():
            parts = geocode_key.split("|", 1)
            if len(parts) != 2:
                geocode_no_match += 1
                continue

            addr_part, postal_part = parts
            lookup_key = f"{normalize_address(addr_part)}|{postal_part.strip()}"

            if lookup_key not in db_by_addr:
                geocode_no_match += 1
                continue

            for loc_row in db_by_addr[lookup_key]:
                if loc_row["latitude"] is not None and loc_row["longitude"] is not None:
                    if (abs(loc_row["latitude"] - lat) < 0.0001
                            and abs(loc_row["longitude"] - lon) < 0.0001):
                        already_set_addr += 1
                        updated_location_ids.add(loc_row["id"])
                        continue

                if not dry_run:
                    await conn.execute(
                        "UPDATE locations SET latitude = $1, longitude = $2 WHERE id = $3",
                        lat, lon, loc_row["id"],
                    )
                updated_by_addr += 1
                updated_location_ids.add(loc_row["id"])

        print(f"Pass 2 (adresse):      {updated_by_addr} oppdatert, "
              f"{already_set_addr} allerede satt, {geocode_no_match} uten match i DB")

        # ── Pass 3: Google Maps Geocoding for remaining locations ─────────────
        geocoded_api = 0
        geocode_failed = 0

        if do_geocode:
            # Find locations still without coordinates
            remaining_locs = [
                r for r in db_locations
                if r["id"] not in updated_location_ids
                and (r["latitude"] is None or r["longitude"] is None)
                and r["address"] and r["address"] != "Ukjent adresse"
                and r["postal_code"] and r["postal_code"] != "0000"
            ]

            print(f"\nPass 3 (Google Maps API): {len(remaining_locs)} lokasjoner a geokode")

            for i, loc_row in enumerate(remaining_locs, 1):
                addr = loc_row["address"]
                postal = loc_row["postal_code"]

                result = geocode_address(addr, postal, api_key)

                if result:
                    lat, lon = result
                    if not dry_run:
                        await conn.execute(
                            "UPDATE locations SET latitude = $1, longitude = $2 WHERE id = $3",
                            lat, lon, loc_row["id"],
                        )
                    geocoded_api += 1
                    if i <= 5 or i % 50 == 0:
                        print(f"  [{i}/{len(remaining_locs)}] {addr}, {postal} -> "
                              f"{lat:.6f}, {lon:.6f}")
                else:
                    geocode_failed += 1
                    if i <= 5 or geocode_failed <= 10:
                        print(f"  [{i}/{len(remaining_locs)}] FEILET: {addr}, {postal}")

                # Rate limiting: 50 req/s limit, but be conservative
                time.sleep(0.1)

            print(f"Pass 3 (Google Maps):  {geocoded_api} geokodet, "
                  f"{geocode_failed} feilet")
        else:
            # Just count remaining
            remaining_count = sum(
                1 for r in db_locations
                if r["id"] not in updated_location_ids
                and (r["latitude"] is None or r["longitude"] is None)
            )
            if remaining_count > 0:
                print(f"\n  {remaining_count} lokasjoner mangler fortsatt koordinater.")
                print("  Kjor med --geocode for a geokode via Google Maps API.")

        # ── Summary ──────────────────────────────────────────────────────────
        total_locs = await conn.fetchval(
            "SELECT COUNT(*) FROM locations WHERE tenant_id = $1", tenant_id,
        )
        with_coords = await conn.fetchval(
            "SELECT COUNT(*) FROM locations WHERE tenant_id = $1 "
            "AND latitude IS NOT NULL AND longitude IS NOT NULL", tenant_id,
        )
        without_coords = total_locs - with_coords

        print()
        print("=" * 55)
        print("Import av koordinater ferdig!")
        print("=" * 55)
        print()
        print("  Kilder:")
        for region_name, count in excel_counts.items():
            print(f"    Excel {region_name}: {count} rader med lat/lon")
        for region_name, count in geocode_counts.items():
            print(f"    Geocode {region_name}: {count} adresser med lat/lon")
        if do_geocode:
            print(f"    Google Maps API: {geocoded_api} geokodet, {geocode_failed} feilet")
        print()
        print("  Resultater:")
        print(f"    Oppdatert via ticket number:   {updated_by_ticket}")
        print(f"    Oppdatert via adresse-match:   {updated_by_addr}")
        print(f"    Oppdatert via Google Maps:     {geocoded_api}")
        print(f"    Allerede satt (uendret):       {already_set_ticket + already_set_addr}")
        print()
        print("  Lokasjoner i v2:")
        print(f"    Totalt:                        {total_locs}")
        print(f"    Med koordinater:               {with_coords}")
        print(f"    Uten koordinater:              {without_coords}")

    finally:
        await conn.close()


if __name__ == "__main__":
    asyncio.run(import_coordinates())
