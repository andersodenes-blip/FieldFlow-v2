# Copyright (c) 2026 Anders Ødenes. All rights reserved.
"""Import all jobs from the Dynamics master Excel export — asyncpg direct."""
import asyncio
import os
import sys
import uuid
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

# Add project root to path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# Load .env BEFORE any app imports
from dotenv import load_dotenv

load_dotenv(PROJECT_ROOT / ".env", override=True)
DATABASE_URL = os.environ["DATABASE_URL"].replace("postgresql+asyncpg://", "postgresql://")

try:
    import openpyxl
except ImportError:
    print("openpyxl er ikke installert. Kjør: pip install openpyxl")
    sys.exit(1)

import asyncpg

EXCEL_FILE = PROJECT_ROOT / "data" / "master.xlsx"
SHEET_NAME = "Årskontroller Hedengren Master"

# Dynamics region name → v2 region name
REGION_MAPPING = {
    "Sør-Norge": "Stavanger",
    "Oslo": "Oslo",
    "Drammen": "Drammen",
    "Bergen": "Bergen",
    "Østfold": "Østfold",
    "Innlandet": "Innlandet",
    "Midt-Norge": "Innlandet",
}


async def import_master():
    print(f"Database: {DATABASE_URL[:40]}...")

    # Read Excel
    if not EXCEL_FILE.exists():
        print(f"Excel-fil ikke funnet: {EXCEL_FILE}")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        print(f"Ark '{SHEET_NAME}' ikke funnet. Tilgjengelige: {wb.sheetnames}")
        wb.close()
        return

    now = datetime.now(timezone.utc)

    conn = await asyncpg.connect(DATABASE_URL, statement_cache_size=0)
    try:
        # 1. Find tenant
        row = await conn.fetchrow("SELECT id FROM tenants WHERE slug = $1", "hedengren")
        if not row:
            print("Tenant 'hedengren' not found. Run seed.py first.")
            wb.close()
            return

        tenant_id = row["id"]
        print(f"Tenant: hedengren (id: {tenant_id})")

        # 2. Set RLS context
        await conn.execute(f"SET app.current_tenant = '{tenant_id}'")

        # 3. Load existing regions
        rows = await conn.fetch("SELECT id, name FROM regions WHERE tenant_id = $1", tenant_id)
        region_map = {r["name"]: r["id"] for r in rows}

        # Create missing regions
        needed_regions = set(REGION_MAPPING.values()) - set(region_map.keys())
        for region_name in sorted(needed_regions):
            rid = uuid.uuid4()
            await conn.execute(
                "INSERT INTO regions (id, tenant_id, name, city, created_at, updated_at) "
                "VALUES ($1, $2, $3, $4, $5, $6)",
                rid, tenant_id, region_name, region_name, now, now,
            )
            region_map[region_name] = rid
            print(f"  Opprettet region: {region_name}")

        print(f"Regioner: {sorted(region_map.keys())}")

        # 4. Find or create customer per region
        customer_map = {}
        for region_name in region_map:
            cust_name = f"{region_name} Kunder"
            row = await conn.fetchrow(
                "SELECT id FROM customers WHERE tenant_id = $1 AND name = $2",
                tenant_id, cust_name,
            )
            if row:
                cust_id = row["id"]
            else:
                cust_id = uuid.uuid4()
                await conn.execute(
                    "INSERT INTO customers (id, tenant_id, name, created_at, updated_at) "
                    "VALUES ($1, $2, $3, $4, $5)",
                    cust_id, tenant_id, cust_name, now, now,
                )
                print(f"  Opprettet kunde: {cust_name}")
            customer_map[region_name] = cust_id

        # 5. Load existing locations and jobs by external_id
        rows = await conn.fetch(
            "SELECT external_id, id FROM locations WHERE tenant_id = $1 AND external_id IS NOT NULL",
            tenant_id,
        )
        location_map = {r["external_id"]: r["id"] for r in rows}

        rows = await conn.fetch(
            "SELECT external_id FROM jobs WHERE tenant_id = $1 AND external_id IS NOT NULL",
            tenant_id,
        )
        existing_job_ext_ids = {r["external_id"] for r in rows}

        # 6. Load existing service contracts by location_id
        rows = await conn.fetch(
            "SELECT location_id, id FROM service_contracts WHERE tenant_id = $1",
            tenant_id,
        )
        contract_map = {r["location_id"]: r["id"] for r in rows}

        # 7. Process rows
        region_stats = defaultdict(lambda: {
            "locations": 0, "contracts": 0, "jobs": 0, "skipped": 0,
        })
        rows_no_data = 0
        rows_unknown_region = 0
        row_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 11:
                rows_no_data += 1
                continue

            ticket_number = str(row[3]).strip() if row[3] else ""
            site_number = str(row[4]).strip() if row[4] else ""
            system_type = str(row[6]).strip() if row[6] else ""
            service_region = str(row[7]).strip() if row[7] else ""
            address = str(row[9]).strip() if row[9] else ""
            postal_code = str(row[10]).strip() if row[10] else ""

            if not ticket_number or not site_number:
                rows_no_data += 1
                continue

            v2_region = REGION_MAPPING.get(service_region)
            if not v2_region:
                rows_unknown_region += 1
                continue

            region_id = region_map[v2_region]
            customer_id = customer_map[v2_region]
            stats = region_stats[v2_region]

            # a. Find or create Location by Site Number
            if site_number in location_map:
                location_id = location_map[site_number]
            else:
                location_id = uuid.uuid4()
                await conn.execute(
                    "INSERT INTO locations (id, tenant_id, customer_id, address, city, postal_code, external_id, created_at, updated_at) "
                    "VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)",
                    location_id, tenant_id, customer_id,
                    address or "Ukjent adresse", v2_region,
                    postal_code or "0000", site_number, now, now,
                )
                location_map[site_number] = location_id
                stats["locations"] += 1

            # b. Find or create ServiceContract for location
            if location_id not in contract_map:
                contract_id = uuid.uuid4()
                await conn.execute(
                    "INSERT INTO service_contracts (id, tenant_id, location_id, service_type, interval_months, next_due_date, is_active, created_at, updated_at) "
                    "VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)",
                    contract_id, tenant_id, location_id,
                    system_type or "Årskontroll", 12,
                    date(2027, 1, 1), True, now, now,
                )
                contract_map[location_id] = contract_id
                stats["contracts"] += 1

            contract_id = contract_map[location_id]

            # c. Create Job if not already imported
            if ticket_number in existing_job_ext_ids:
                stats["skipped"] += 1
                continue

            job_id = uuid.uuid4()
            title = f"{system_type} — {site_number}" if system_type else f"Årskontroll — {site_number}"
            await conn.execute(
                "INSERT INTO jobs (id, tenant_id, service_contract_id, title, description, status, external_id, created_at, updated_at) "
                "VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)",
                job_id, tenant_id, contract_id,
                title, f"Ticket: {ticket_number}, Site: {site_number}",
                "unscheduled", ticket_number, now, now,
            )
            existing_job_ext_ids.add(ticket_number)
            stats["jobs"] += 1

            row_count += 1
            if row_count % 100 == 0:
                print(f"  ... {row_count} rader behandlet")

    finally:
        await conn.close()

    wb.close()

    # Print stats
    print(f"\n{'='*50}")
    print("Import ferdig!")
    print(f"{'='*50}")

    total = {"locations": 0, "contracts": 0, "jobs": 0, "skipped": 0}
    for region_name in sorted(region_stats.keys()):
        s = region_stats[region_name]
        print(f"\n  {region_name}:")
        print(f"    Lokasjoner opprettet: {s['locations']}")
        print(f"    Serviceavtaler opprettet: {s['contracts']}")
        print(f"    Jobber opprettet: {s['jobs']}")
        print(f"    Jobber hoppet over: {s['skipped']}")
        for k in total:
            total[k] += s[k]

    print(f"\n  TOTALT:")
    print(f"    Lokasjoner opprettet: {total['locations']}")
    print(f"    Serviceavtaler opprettet: {total['contracts']}")
    print(f"    Jobber opprettet: {total['jobs']}")
    print(f"    Jobber hoppet over: {total['skipped']}")
    print(f"    Rader uten data: {rows_no_data}")
    if rows_unknown_region:
        print(f"    Ukjent region (hoppet over): {rows_unknown_region}")


if __name__ == "__main__":
    asyncio.run(import_master())
