"""Inspect Master.xlsx structure."""
import openpyxl

wb = openpyxl.load_workbook(r"C:\FieldFlow-v2\data\Master.xlsx", read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)
for s in wb.sheetnames:
    ws = wb[s]
    print(f"\n=== {s}: max_row={ws.max_row}, max_column={ws.max_column} ===")
    rows = list(ws.iter_rows(max_row=3, values_only=True))
    if rows:
        for i, h in enumerate(rows[0]):
            val = str(h)[:80] if h else None
            print(f"  [{i:2d}] {val}")
        if len(rows) > 1:
            print(f"\n  Rad 2: {rows[1][:12]}")
        if len(rows) > 2:
            print(f"  Rad 3: {rows[2][:12]}")
wb.close()
