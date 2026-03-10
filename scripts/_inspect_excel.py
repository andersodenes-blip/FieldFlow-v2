"""Temporary script to inspect Excel column names."""
import openpyxl
import sys

FILES = {
    "bergen": (
        r"C:\Users\anders.odenes\OneDrive - Oy Hedengren Ab"
        r"\Työpöytä\FieldFlow\data\bergen\Årskontroller Bergen.xlsx"
    ),
    "oslo": (
        r"C:\Users\anders.odenes\OneDrive - Oy Hedengren Ab"
        r"\Työpöytä\FieldFlow\data\oslo\Årskontroller Oslo.xlsx"
    ),
    "stavanger": (
        r"C:\Users\anders.odenes\OneDrive - Oy Hedengren Ab"
        r"\Työpöytä\FieldFlow\data\stavanger\Årskontroller Stavanger.xlsx"
    ),
    "drammen": (
        r"C:\Users\anders.odenes\OneDrive - Oy Hedengren Ab"
        r"\Työpöytä\FieldFlow\data\drammen\Årskontroller Drammen.xlsx"
    ),
    "master": r"C:\FieldFlow-v2\data\Master.xlsx",
    "master-stavanger": (
        r"C:\FieldFlow-v2\data\stavanger"
        r"\Årskontroller Hedengren Master 3-9-2026 1-09-24 PM.xlsx"
    ),
}

arg = sys.argv[1] if len(sys.argv) > 1 else "bergen"
filepath = FILES.get(arg.lower(), arg)

wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(max_row=3, values_only=True))
    print(f"\n=== Sheet: '{sheet_name}' ===")
    if rows:
        headers = rows[0]
        for i, h in enumerate(headers):
            print(f"  [{i:2d}] {h}")
        if len(rows) > 1:
            print(f"\n  Rad 2 (sample): {rows[1][:10]}...")
wb.close()
