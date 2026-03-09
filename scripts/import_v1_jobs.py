# Copyright (c) 2026 Anders Ødenes. All rights reserved.
"""Import Sites and Tickets from Dynamics Excel export (Stavanger)."""
import asyncio
import sys
import uuid
from datetime import date
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    import openpyxl
except ImportError:
    print("openpyxl er ikke installert. Kjør: pip install openpyxl")
    sys.exit(1)

from sqlalchemy import text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.config import settings
from app.models.customer import Customer
from app.models.job import Job, JobStatus
from app.models.location import Location
from app.models.service_contract import ServiceContract

PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_FILE = (
    PROJECT_ROOT
    / "data"
    / "stavanger"
    / "Årskontroller Hedengren Master 3-9-2026 1-09-24 PM.xlsx"
)
SHEET_NAME = "Årskontroller Hedengren Master"
REGION_NAME = "Stavanger"
CUSTOMER_NAME = "Stavanger Kunder"


async def import_jobs():
    engine = create_async_engine(settings.DATABASE_URL)
    session_factory = async_sessionmaker(engine, expire_on_commit=False)

    # Read Excel
    if not EXCEL_FILE.exists():
        print(f"Excel-fil ikke funnet: {EXCEL_FILE}")
        await engine.dispose()
        return

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        print(f"Ark '{SHEET_NAME}' ikke funnet. Tilgjengelige: {wb.sheetnames}")
        wb.close()
        await engine.dispose()
        return

    async with session_factory() as db:
        # 1. Find tenant
        result = await db.execute(
            text("SELECT id FROM tenants WHERE slug = 'hedengren'")
        )
        tenant_id = result.scalar_one_or_none()
        if not tenant_id:
            print("Tenant 'hedengren' not found. Run seed.py first.")
            wb.close()
            await engine.dispose()
            return

        print(f"Tenant: hedengren (id: {tenant_id})")

        # 2. Set RLS context
        await db.execute(text(f"SET app.current_tenant = '{tenant_id}'"))

        # 3. Find region
        result = await db.execute(
            text("SELECT id FROM regions WHERE tenant_id = :tid AND name = :name"),
            {"tid": tenant_id, "name": REGION_NAME},
        )
        region_id = result.scalar_one_or_none()
        if not region_id:
            print(f"Region '{REGION_NAME}' ikke funnet.")
            wb.close()
            await engine.dispose()
            return

        print(f"Region: {REGION_NAME} (id: {region_id})")

        # 4. Find or create customer
        result = await db.execute(
            text("SELECT id FROM customers WHERE tenant_id = :tid AND name = :name"),
            {"tid": tenant_id, "name": CUSTOMER_NAME},
        )
        customer_id = result.scalar_one_or_none()
        if not customer_id:
            customer = Customer(
                id=uuid.uuid4(),
                tenant_id=tenant_id,
                name=CUSTOMER_NAME,
            )
            db.add(customer)
            await db.flush()
            customer_id = customer.id
            print(f"Opprettet kunde: {CUSTOMER_NAME} (id: {customer_id})")
        else:
            print(f"Bruker eksisterende kunde: {CUSTOMER_NAME} (id: {customer_id})")

        # 5. Load existing locations and jobs by external_id
        result = await db.execute(
            text("SELECT external_id, id FROM locations WHERE tenant_id = :tid AND external_id IS NOT NULL"),
            {"tid": tenant_id},
        )
        location_map = {row.external_id: row.id for row in result.fetchall()}

        result = await db.execute(
            text("SELECT external_id FROM jobs WHERE tenant_id = :tid AND external_id IS NOT NULL"),
            {"tid": tenant_id},
        )
        existing_job_ext_ids = {row.external_id for row in result.fetchall()}

        # 6. Load existing service contracts by location_id
        result = await db.execute(
            text("SELECT location_id, id FROM service_contracts WHERE tenant_id = :tid"),
            {"tid": tenant_id},
        )
        contract_map = {row.location_id: row.id for row in result.fetchall()}

        # 7. Process rows (skip header row)
        stats = {
            "locations_created": 0,
            "contracts_created": 0,
            "jobs_created": 0,
            "jobs_skipped": 0,
            "rows_skipped_no_data": 0,
        }

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Columns: D=3, E=4, G=6, J=9, K=10 (0-indexed)
            if len(row) < 11:
                stats["rows_skipped_no_data"] += 1
                continue

            ticket_number = str(row[3]).strip() if row[3] else ""
            site_number = str(row[4]).strip() if row[4] else ""
            system_type = str(row[6]).strip() if row[6] else ""
            address = str(row[9]).strip() if row[9] else ""
            postal_code = str(row[10]).strip() if row[10] else ""

            if not ticket_number or not site_number:
                stats["rows_skipped_no_data"] += 1
                continue

            # a. Find or create Location by Site Number
            if site_number in location_map:
                location_id = location_map[site_number]
            else:
                location = Location(
                    id=uuid.uuid4(),
                    tenant_id=tenant_id,
                    customer_id=customer_id,
                    address=address or "Ukjent adresse",
                    city="Stavanger",
                    postal_code=postal_code or "0000",
                    external_id=site_number,
                )
                db.add(location)
                await db.flush()
                location_id = location.id
                location_map[site_number] = location_id
                stats["locations_created"] += 1
                print(f"  Lokasjon: {site_number} — {address} {postal_code}")

            # Ensure a service contract exists for the location
            if location_id not in contract_map:
                contract = ServiceContract(
                    id=uuid.uuid4(),
                    tenant_id=tenant_id,
                    location_id=location_id,
                    service_type=system_type or "Årskontroll",
                    interval_months=12,
                    next_due_date=date(2027, 1, 1),
                    is_active=True,
                )
                db.add(contract)
                await db.flush()
                contract_map[location_id] = contract.id
                stats["contracts_created"] += 1

            contract_id = contract_map[location_id]

            # b. Create Job if not already imported
            if ticket_number in existing_job_ext_ids:
                stats["jobs_skipped"] += 1
                continue

            job = Job(
                id=uuid.uuid4(),
                tenant_id=tenant_id,
                service_contract_id=contract_id,
                title=f"{system_type} — {site_number}" if system_type else f"Årskontroll — {site_number}",
                description=f"Ticket: {ticket_number}, Site: {site_number}",
                status=JobStatus.unscheduled,
                external_id=ticket_number,
            )
            db.add(job)
            existing_job_ext_ids.add(ticket_number)
            stats["jobs_created"] += 1

        await db.commit()

    wb.close()
    await engine.dispose()

    print(f"\nFerdig!")
    print(f"  Lokasjoner opprettet: {stats['locations_created']}")
    print(f"  Serviceavtaler opprettet: {stats['contracts_created']}")
    print(f"  Jobber opprettet: {stats['jobs_created']}")
    print(f"  Jobber hoppet over (finnes): {stats['jobs_skipped']}")
    print(f"  Rader uten data: {stats['rows_skipped_no_data']}")


if __name__ == "__main__":
    asyncio.run(import_jobs())
